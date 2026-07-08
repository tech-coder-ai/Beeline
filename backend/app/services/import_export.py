"""CSV / Excel metadata import with preview-then-commit via the approval queue."""
from __future__ import annotations

import csv
import io
import json

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import ValidationFailed
from app.models.catalog import CatalogColumn, CatalogDatabase, CatalogTable
from app.models.governance import ApprovalItem

# Expected columns (case-insensitive; extras ignored):
#   database, table, column (optional), description, glossary, owner, tags, classification
SUPPORTED_FIELDS = {"database", "table", "column", "description", "glossary",
                    "owner", "tags", "classification", "business_rules"}


class MetadataImportService:
    def parse(self, filename: str, content: bytes) -> list[dict]:
        if filename.lower().endswith((".xlsx", ".xlsm")):
            rows = self._parse_excel(content)
        elif filename.lower().endswith(".csv"):
            rows = self._parse_csv(content)
        else:
            raise ValidationFailed("Only .csv and .xlsx files are supported")
        if not rows:
            raise ValidationFailed("The file contains no data rows")
        return rows

    @staticmethod
    def _parse_csv(content: bytes) -> list[dict]:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [
            {k.strip().lower(): (v or "").strip() for k, v in row.items() if k}
            for row in reader
        ]

    @staticmethod
    def _parse_excel(content: bytes) -> list[dict]:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(h or "").strip().lower() for h in next(rows_iter, [])]
        return [
            {header[i]: str(cell or "").strip() for i, cell in enumerate(row) if i < len(header)}
            for row in rows_iter
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

    async def preview(self, db: AsyncSession, rows: list[dict]) -> dict:
        """Match rows against the catalog and report what would change."""
        matched, unmatched, changes = [], [], []
        for index, row in enumerate(rows):
            db_name = row.get("database", "")
            table_name = row.get("table", "")
            column_name = row.get("column", "")
            if not table_name:
                unmatched.append({"row": index + 1, "reason": "missing table name"})
                continue
            table = await self._find_table(db, db_name, table_name)
            if table is None:
                unmatched.append({"row": index + 1, "reason": f"table '{db_name}.{table_name}' not in catalog"})
                continue
            target_column = None
            if column_name:
                target_column = next((c for c in table.columns if c.name == column_name), None)
                if target_column is None:
                    unmatched.append({"row": index + 1,
                                      "reason": f"column '{column_name}' not in {table_name}"})
                    continue
            matched.append(index + 1)
            changes.extend(self._diff(row, table, target_column))
        return {"matched_rows": len(matched), "unmatched": unmatched, "changes": changes}

    async def commit(self, db: AsyncSession, rows: list[dict]) -> dict:
        """Queue all changes as approval items (source=import)."""
        preview = await self.preview(db, rows)
        queued = 0
        for change in preview["changes"]:
            db.add(ApprovalItem(
                entity_type=change["entity_type"],
                entity_id=change["entity_id"],
                entity_label=change["label"],
                field=change["field"],
                current_value=change["current"],
                proposed_value=change["proposed"],
                proposed_payload=change.get("payload"),
                source="import",
                confidence=1.0,
                rationale="Imported from uploaded file",
            ))
            queued += 1
        return {**preview, "queued_for_approval": queued}

    async def _find_table(self, db: AsyncSession, db_name: str, table_name: str) -> CatalogTable | None:
        stmt = (
            select(CatalogTable)
            .join(CatalogDatabase, CatalogTable.database_id == CatalogDatabase.id)
            .options(selectinload(CatalogTable.columns), selectinload(CatalogTable.database))
            .where(CatalogTable.name == table_name)
        )
        if db_name:
            stmt = stmt.where(CatalogDatabase.name == db_name)
        return (await db.execute(stmt)).scalars().first()

    @staticmethod
    def _diff(row: dict, table: CatalogTable, column: CatalogColumn | None) -> list[dict]:
        changes = []
        label = f"{table.database.name}.{table.name}" + (f".{column.name}" if column else "")

        def add(entity_type: str, entity_id: str, field: str, current, proposed, payload=None):
            if proposed and str(proposed) != str(current or ""):
                changes.append({
                    "entity_type": entity_type, "entity_id": entity_id, "label": label,
                    "field": field, "current": current, "proposed": proposed, "payload": payload,
                })

        if column is not None:
            add("column_description", column.id, "description",
                column.description, row.get("description"))
        else:
            add("table_description", table.id, "description",
                table.description, row.get("description"))
            add("classification", table.id, "classification",
                table.classification, row.get("classification"))
            if row.get("tags"):
                proposed_tags = sorted({t.strip() for t in row["tags"].split(",") if t.strip()})
                add("tag", table.id, "tags", json.dumps(table.tags or []),
                    json.dumps(proposed_tags))
        if row.get("glossary"):
            add("glossary_term", table.id, "glossary", None, row["glossary"])
        return changes


import_service = MetadataImportService()
